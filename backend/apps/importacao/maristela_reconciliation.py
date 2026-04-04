from __future__ import annotations

import csv
import json
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.conf import settings
from django.core.files.base import ContentFile
from django.utils import timezone
from openpyxl import load_workbook

from apps.accounts.models import User
from apps.associados.models import Associado, only_digits
from apps.contratos.competencia import propagate_competencia_status
from apps.contratos.cycle_rebuild import rebuild_contract_cycle_state
from apps.contratos.models import Contrato, Parcela
from apps.importacao.manual_payment_flags import MANUAL_PAYMENT_KIND_MARISTELA_IN_CYCLE
from apps.importacao.matching import normalize_matricula
from apps.importacao.models import PagamentoMensalidade
from apps.tesouraria.models import BaixaManual

PERIOD_REFERENCE_MAP: dict[str, date] = {
    "outubro/25": date(2025, 10, 1),
    "dezembro/25": date(2025, 12, 1),
    "janeiro/26": date(2026, 1, 1),
    "fevereiro/26": date(2026, 2, 1),
    "março/26": date(2026, 3, 1),
}
PERIOD_REFERENCES = tuple(PERIOD_REFERENCE_MAP.values())
PERIOD_START = min(PERIOD_REFERENCES)
PERIOD_END = max(PERIOD_REFERENCES)

SYSTEM_ONLY_FIELDS = [
    "associado_id",
    "cpf_cnpj",
    "nome_completo",
    "status",
    "matricula",
    "matricula_orgao",
    "total_contratos",
    "contratos_ativos",
    "referencias_no_periodo",
]
UNMATCHED_FIELDS = [
    "row_number",
    "reason",
    "cpf_cnpj",
    "nome_planilha",
    "matricula_planilha",
    "detail",
]
EXCEPTION_FIELDS = [
    "row_number",
    "cpf_cnpj",
    "nome_planilha",
    "matricula_planilha",
    "competencia",
    "reason",
    "detail",
]
CORRECTION_FIELDS = [
    "row_number",
    "cpf_cnpj",
    "nome_completo",
    "associado_id",
    "contrato_id",
    "contrato_codigo",
    "competencia",
    "entity",
    "entity_id",
    "action",
    "reason",
    "before",
    "after",
]
SHEET_SOURCE_PATH = "conciliacao/planilha_manual_maristela.xlsx"


def _fold_text(value: str | None) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    collapsed = "".join(char for char in normalized if not unicodedata.combining(char))
    return " ".join(collapsed.casefold().split())


def _append_note(base: str, note: str) -> str:
    if not base:
        return note
    if note in base:
        return base
    return f"{base}\n{note}"


def _normalize_sheet_document(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float):
        rendered = str(int(value))
    else:
        rendered = str(value).strip()
    digits = only_digits(rendered)
    if not digits:
        return ""
    if len(digits) <= 11:
        return digits.zfill(11)
    return digits


def _normalize_sheet_matricula(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float):
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _decimal_or_none(value: object) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _reference_label(value: date | None) -> str | None:
    if value is None:
        return None
    return value.isoformat()


def _default_report_dir() -> Path:
    timestamp = datetime.now().strftime("%Y%m%dT%H%M%S")
    return (
        Path(settings.BASE_DIR)
        / "media"
        / "relatorios"
        / "legacy_import"
        / f"maristela_reconciliation_{timestamp}"
    )


def _default_paid_date(referencia: date) -> date:
    return referencia


def _default_manual_paid_at(referencia: date):
    return timezone.make_aware(datetime.combine(referencia, time(12, 0)))


@dataclass(frozen=True)
class MaristelaRow:
    row_number: int
    cpf_cnpj: str
    nome_planilha: str
    mensalidade: Decimal | None
    matricula_planilha: str
    competencias: dict[date, str]


@dataclass(frozen=True)
class CellTarget:
    kind: str
    associado_status: str
    pagamento_status_code: str
    descricao: str


@dataclass(frozen=True)
class MatchResult:
    associado: Associado | None
    reason: str | None
    detail: dict[str, Any]


@dataclass
class AssociatedFinanceContext:
    parcelas_by_reference: dict[date, list[Parcela]]
    pagamentos_by_reference: dict[date, list[PagamentoMensalidade]]
    baixa_by_parcela_id: dict[int, BaixaManual]


class MaristelaReconciliationRunner:
    def __init__(
        self,
        *,
        file_path: str | Path,
        execute: bool,
        actor: User | None = None,
    ) -> None:
        self.file_path = Path(file_path).expanduser().resolve()
        self.execute = execute
        self.actor = actor
        self.generated_at = timezone.now()
        self._all_associados = list(
            Associado.objects.all().order_by("id")
        )
        self._cpf_index: dict[str, list[Associado]] = defaultdict(list)
        self._matricula_index: dict[str, list[Associado]] = defaultdict(list)
        for associado in self._all_associados:
            cpf = only_digits(associado.cpf_cnpj)
            if cpf:
                self._cpf_index[cpf].append(associado)
            for raw_matricula in [associado.matricula, associado.matricula_orgao]:
                normalized = normalize_matricula(raw_matricula)
                if normalized:
                    self._matricula_index[normalized].append(associado)

    def run(self) -> dict[str, Any]:
        rows = self._load_rows()
        sheet_cpfs = {row.cpf_cnpj for row in rows if row.cpf_cnpj}
        sheet_matriculas = {
            normalize_matricula(row.matricula_planilha)
            for row in rows
            if normalize_matricula(row.matricula_planilha)
        }

        unmatched_rows: list[dict[str, Any]] = []
        exceptions: list[dict[str, Any]] = []
        corrections: list[dict[str, Any]] = []
        matched_associado_ids: set[int] = set()
        affected_contract_ids: set[int] = set()
        stats = Counter()
        by_competencia: dict[str, Counter[str]] = defaultdict(Counter)

        for row in rows:
            stats["rows_total"] += 1
            match = self._match_associado(row)
            if match.associado is None:
                stats["rows_without_unique_match"] += 1
                unmatched_rows.append(
                    {
                        "row_number": row.row_number,
                        "reason": match.reason,
                        "cpf_cnpj": row.cpf_cnpj,
                        "nome_planilha": row.nome_planilha,
                        "matricula_planilha": row.matricula_planilha,
                        "detail": match.detail,
                    }
                )
                continue

            associado = match.associado
            matched_associado_ids.add(associado.id)
            context = self._load_associado_context(associado)
            row_target_status = self._derive_global_status(row)
            touched_contracts: set[int] = set()

            for referencia, raw_value in sorted(row.competencias.items(), key=lambda item: item[0]):
                stats["cells_total"] += 1
                try:
                    target = self._classify_cell(raw_value)
                except ValueError as exc:
                    stats["cells_with_unsupported_value"] += 1
                    exceptions.append(
                        self._build_exception_row(
                            row,
                            referencia,
                            "unsupported_value",
                            {"raw_value": raw_value, "error": str(exc)},
                        )
                    )
                    continue

                parcelas = context.parcelas_by_reference.get(referencia, [])
                if not parcelas:
                    stats["cells_without_parcela"] += 1
                    exceptions.append(
                        self._build_exception_row(
                            row,
                            referencia,
                            "parcela_not_found",
                            {"raw_value": raw_value},
                        )
                    )
                    continue
                if len(parcelas) > 1:
                    stats["cells_with_parcela_conflict"] += 1
                    exceptions.append(
                        self._build_exception_row(
                            row,
                            referencia,
                            "parcela_conflict",
                            {
                                "raw_value": raw_value,
                                "parcelas": [
                                    self._serialize_parcela(parcela) for parcela in parcelas
                                ],
                            },
                        )
                    )
                    continue

                parcela = parcelas[0]
                pagamentos = list(context.pagamentos_by_reference.get(referencia, []))
                baixa = context.baixa_by_parcela_id.get(parcela.id)

                month_actions = self._reconcile_reference(
                    row=row,
                    associado=associado,
                    referencia=referencia,
                    raw_value=raw_value,
                    target=target,
                    parcela=parcela,
                    pagamentos=pagamentos,
                    baixa=baixa,
                )
                if month_actions:
                    touched_contracts.add(parcela.ciclo.contrato_id)
                    stats["cells_with_divergence"] += 1
                    stats["corrections_total"] += len(month_actions)
                    by_competencia[referencia.isoformat()]["corrections"] += len(month_actions)
                    corrections.extend(month_actions)

            if row_target_status and row_target_status != associado.status:
                correction = self._build_correction(
                    row=row,
                    associado=associado,
                    contrato=None,
                    referencia=None,
                    entity="associado",
                    entity_id=associado.id,
                    action="update_status",
                    reason="derived_from_sheet",
                    before={
                        "status": associado.status,
                        "observacao": associado.observacao,
                    },
                    after={
                        "status": row_target_status,
                        "observacao": _append_note(
                            associado.observacao,
                            self._build_associado_note(row_target_status),
                        ),
                    },
                )
                corrections.append(correction)
                stats["associado_status_updates"] += 1
                if self.execute:
                    associado.status = row_target_status
                    associado.observacao = correction["after"]["observacao"]
                    associado.save(update_fields=["status", "observacao", "updated_at"])

            if touched_contracts:
                affected_contract_ids.update(touched_contracts)
                if self.execute:
                    for contrato_id in sorted(touched_contracts):
                        contrato = Contrato.objects.select_related("associado").get(pk=contrato_id)
                        rebuild_contract_cycle_state(contrato, execute=True)
                corrections.extend(
                    [
                        self._build_correction(
                            row=row,
                            associado=associado,
                            contrato=Contrato.objects.only("id", "codigo").get(pk=contrato_id),
                            referencia=None,
                            entity="contrato",
                            entity_id=contrato_id,
                            action="rebuild_cycle_state",
                            reason="post_reconciliation",
                            before={},
                            after={"execute": self.execute},
                        )
                        for contrato_id in sorted(touched_contracts)
                    ]
                )

        system_only = self._list_system_outside_sheet(sheet_cpfs, sheet_matriculas)
        prioritized_system_only = [
            row for row in system_only if row["referencias_no_periodo"]
        ]

        post_process: dict[str, Any] = {}
        if self.execute:
            post_process = self._run_post_process()

        summary = {
            "generated_at": self.generated_at.isoformat(),
            "mode": "execute" if self.execute else "dry-run",
            "file": str(self.file_path),
            "period_references": [_reference_label(value) for value in PERIOD_REFERENCES],
            "rows_loaded": len(rows),
            "matched_associados": len(matched_associado_ids),
            "planilha_sem_match": len(unmatched_rows),
            "excecoes_conciliacao": len(exceptions),
            "correcoes_planejadas": len(corrections),
            "affected_contracts": len(affected_contract_ids),
            "stats": dict(stats),
            "by_competencia": {
                key: dict(value) for key, value in sorted(by_competencia.items())
            },
            "post_process": post_process,
        }

        return {
            "summary": summary,
            "corrections": corrections,
            "planilha_sem_match": unmatched_rows,
            "excecoes_conciliacao": exceptions,
            "sistema_fora_da_planilha": system_only,
            "sistema_fora_da_planilha_priorizados": prioritized_system_only,
        }

    def _load_rows(self) -> list[MaristelaRow]:
        if not self.file_path.exists():
            raise FileNotFoundError(f"Planilha não encontrada: {self.file_path}")

        workbook = load_workbook(self.file_path, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        header_index: dict[int, str] = {}
        for column in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, column).value
            header = str(cell_value or "").strip()
            if header:
                header_index[column] = header

        required_headers = {"CPF", "NOME", "MENSALIDADE", "MATRICULA"}
        normalized_headers = {_fold_text(value).replace(" ", ""): key for key, value in header_index.items()}
        missing = [
            header
            for header in required_headers
            if _fold_text(header).replace(" ", "") not in normalized_headers
        ]
        if missing:
            raise ValueError(
                f"Cabeçalhos obrigatórios ausentes na planilha: {', '.join(sorted(missing))}"
            )

        rows: list[MaristelaRow] = []
        for row_number in range(2, sheet.max_row + 1):
            row_values = {
                header_index[column]: sheet.cell(row_number, column).value
                for column in header_index
            }
            cpf_cnpj = _normalize_sheet_document(row_values.get("CPF"))
            nome_planilha = str(row_values.get("NOME") or row_values.get(" NOME") or "").strip()
            mensalidade = _decimal_or_none(row_values.get("MENSALIDADE"))
            matricula_planilha = _normalize_sheet_matricula(row_values.get("MATRICULA"))

            competencias: dict[date, str] = {}
            for header, referencia in PERIOD_REFERENCE_MAP.items():
                raw_value = row_values.get(header)
                if raw_value in (None, ""):
                    continue
                rendered = " ".join(str(raw_value).split())
                if rendered:
                    competencias[referencia] = rendered

            if not any([cpf_cnpj, nome_planilha, matricula_planilha, competencias]):
                continue

            rows.append(
                MaristelaRow(
                    row_number=row_number,
                    cpf_cnpj=cpf_cnpj,
                    nome_planilha=nome_planilha,
                    mensalidade=mensalidade,
                    matricula_planilha=matricula_planilha,
                    competencias=competencias,
                )
            )
        return rows

    def _match_associado(self, row: MaristelaRow) -> MatchResult:
        cpf_matches = self._cpf_index.get(row.cpf_cnpj, []) if row.cpf_cnpj else []
        matricula_key = normalize_matricula(row.matricula_planilha)
        matricula_matches = self._matricula_index.get(matricula_key, []) if matricula_key else []

        unique_cpf = cpf_matches[0] if len(cpf_matches) == 1 else None
        unique_matricula = matricula_matches[0] if len(matricula_matches) == 1 else None

        if len(cpf_matches) > 1:
            return MatchResult(
                associado=None,
                reason="ambiguous_cpf",
                detail={"associado_ids": [item.id for item in cpf_matches]},
            )
        if len(matricula_matches) > 1:
            return MatchResult(
                associado=None,
                reason="ambiguous_matricula",
                detail={"associado_ids": [item.id for item in matricula_matches]},
            )
        if unique_cpf and unique_matricula and unique_cpf.id != unique_matricula.id:
            return MatchResult(
                associado=None,
                reason="cpf_matricula_conflict",
                detail={
                    "cpf_associado_id": unique_cpf.id,
                    "matricula_associado_id": unique_matricula.id,
                },
            )
        if unique_cpf is not None:
            return MatchResult(associado=unique_cpf, reason=None, detail={"via": "cpf"})
        if unique_matricula is not None:
            return MatchResult(
                associado=unique_matricula,
                reason=None,
                detail={"via": "matricula"},
            )
        return MatchResult(
            associado=None,
            reason="no_match",
            detail={
                "cpf_cnpj": row.cpf_cnpj,
                "matricula_planilha": row.matricula_planilha,
            },
        )

    def _load_associado_context(self, associado: Associado) -> AssociatedFinanceContext:
        parcelas_by_reference: dict[date, list[Parcela]] = defaultdict(list)
        parcelas = list(
            Parcela.all_objects.filter(
                associado=associado,
                referencia_mes__in=PERIOD_REFERENCES,
                deleted_at__isnull=True,
            )
            .exclude(status=Parcela.Status.CANCELADO)
            .select_related("ciclo__contrato")
            .order_by("referencia_mes", "ciclo__contrato_id", "numero", "id")
        )
        for parcela in parcelas:
            parcelas_by_reference[parcela.referencia_mes].append(parcela)

        pagamentos_by_reference: dict[date, list[PagamentoMensalidade]] = defaultdict(list)
        pagamentos = list(
            PagamentoMensalidade.objects.filter(
                referencia_month__in=PERIOD_REFERENCES,
                cpf_cnpj=associado.cpf_cnpj,
            )
            .order_by("referencia_month", "id")
        )
        seen_pagamento_ids: set[int] = set()
        for pagamento in pagamentos:
            pagamentos_by_reference[pagamento.referencia_month].append(pagamento)
            seen_pagamento_ids.add(pagamento.id)
        if associados_payments := list(
            PagamentoMensalidade.objects.filter(
                referencia_month__in=PERIOD_REFERENCES,
                associado=associado,
            )
            .exclude(id__in=seen_pagamento_ids)
            .order_by("referencia_month", "id")
        ):
            for pagamento in associados_payments:
                pagamentos_by_reference[pagamento.referencia_month].append(pagamento)

        baixa_by_parcela_id = {
            baixa.parcela_id: baixa
            for baixa in BaixaManual.objects.select_related("parcela")
            .filter(parcela__associado=associado, parcela__referencia_mes__in=PERIOD_REFERENCES)
            .order_by("id")
        }
        return AssociatedFinanceContext(
            parcelas_by_reference=parcelas_by_reference,
            pagamentos_by_reference=pagamentos_by_reference,
            baixa_by_parcela_id=baixa_by_parcela_id,
        )

    def _classify_cell(self, raw_value: str) -> CellTarget:
        folded = _fold_text(raw_value)
        if not folded:
            raise ValueError("Célula vazia não pode ser conciliada.")
        if folded.startswith("averbado cch de ") or folded.startswith("averbado no cch de "):
            return CellTarget(
                kind="descontado",
                associado_status=Associado.Status.ATIVO,
                pagamento_status_code="1",
                descricao="Competência averbada no CCH.",
            )
        if folded == "baixado parcela manual no sistema aguardando cch de marco":
            return CellTarget(
                kind="manual",
                associado_status=Associado.Status.ATIVO,
                pagamento_status_code="M",
                descricao="Competência quitada manualmente aguardando CCH.",
            )
        if folded == "nao lancado por falta de saldo":
            return CellTarget(
                kind="nao_descontado",
                associado_status=Associado.Status.INADIMPLENTE,
                pagamento_status_code="2",
                descricao="Não lançado por falta de saldo.",
            )
        if folded == "nao lancado: outros":
            return CellTarget(
                kind="nao_descontado",
                associado_status=Associado.Status.INADIMPLENTE,
                pagamento_status_code="3",
                descricao="Não lançado por outros motivos.",
            )
        if folded in {"nao lancado: falecimento.", "nao lancado: falecimento"}:
            return CellTarget(
                kind="nao_descontado",
                associado_status=Associado.Status.INATIVO,
                pagamento_status_code="3",
                descricao="Não lançado por falecimento.",
            )
        if folded in {
            "nao lancado: desligamento.",
            "nao lancado: desligamento",
        }:
            return CellTarget(
                kind="nao_descontado",
                associado_status=Associado.Status.INATIVO,
                pagamento_status_code="3",
                descricao="Não lançado por desligamento.",
            )
        raise ValueError(f"Valor não mapeado: {raw_value}")

    def _derive_global_status(self, row: MaristelaRow) -> str | None:
        desired_status = None
        for raw_value in row.competencias.values():
            try:
                target = self._classify_cell(raw_value)
            except ValueError:
                continue
            if target.associado_status == Associado.Status.INATIVO:
                return Associado.Status.INATIVO
            if (
                target.associado_status == Associado.Status.INADIMPLENTE
                and desired_status != Associado.Status.INATIVO
            ):
                desired_status = Associado.Status.INADIMPLENTE
            elif (
                target.associado_status == Associado.Status.ATIVO
                and desired_status is None
            ):
                desired_status = Associado.Status.ATIVO
        return desired_status

    def _reconcile_reference(
        self,
        *,
        row: MaristelaRow,
        associado: Associado,
        referencia: date,
        raw_value: str,
        target: CellTarget,
        parcela: Parcela,
        pagamentos: list[PagamentoMensalidade],
        baixa: BaixaManual | None,
    ) -> list[dict[str, Any]]:
        corrections: list[dict[str, Any]] = []
        base_value = row.mensalidade or parcela.valor

        if target.kind != "manual" and baixa is not None:
            corrections.append(
                self._build_correction(
                    row=row,
                    associado=associado,
                    contrato=parcela.ciclo.contrato,
                    referencia=referencia,
                    entity="baixa_manual",
                    entity_id=baixa.id,
                    action="delete",
                    reason="target_not_manual",
                    before=self._serialize_baixa(baixa),
                    after={},
                )
            )
            if self.execute:
                baixa.delete()

        if target.kind == "descontado":
            corrections.extend(
                self._ensure_regular_paid(
                    row=row,
                    associado=associado,
                    referencia=referencia,
                    target=target,
                    parcela=parcela,
                )
            )
        elif target.kind == "manual":
            corrections.extend(
                self._ensure_manual_paid(
                    row=row,
                    associado=associado,
                    referencia=referencia,
                    target=target,
                    parcela=parcela,
                    baixa=baixa,
                    valor=base_value or parcela.valor,
                )
            )
        else:
            corrections.extend(
                self._ensure_unpaid(
                    row=row,
                    associado=associado,
                    referencia=referencia,
                    target=target,
                    parcela=parcela,
                )
            )

        corrections.extend(
            self._sync_pagamentos(
                row=row,
                associado=associado,
                referencia=referencia,
                target=target,
                pagamentos=pagamentos,
                valor=base_value or parcela.valor,
            )
        )
        return corrections

    def _ensure_regular_paid(
        self,
        *,
        row: MaristelaRow,
        associado: Associado,
        referencia: date,
        target: CellTarget,
        parcela: Parcela,
    ) -> list[dict[str, Any]]:
        if parcela.status == Parcela.Status.LIQUIDADA and parcela.data_pagamento is not None:
            return []

        desired_date = parcela.data_pagamento or _default_paid_date(referencia)
        desired_status = Parcela.Status.DESCONTADO

        if parcela.status == desired_status and parcela.data_pagamento == desired_date:
            return []

        correction = self._build_correction(
            row=row,
            associado=associado,
            contrato=parcela.ciclo.contrato,
            referencia=referencia,
            entity="parcela",
            entity_id=parcela.id,
            action="update_status",
            reason="sheet_paid",
            before=self._serialize_parcela(parcela),
            after={
                **self._serialize_parcela(parcela),
                "status": desired_status,
                "data_pagamento": _reference_label(desired_date),
                "observacao": parcela.observacao,
            },
        )
        if self.execute:
            parcela.status = desired_status
            parcela.data_pagamento = desired_date
            parcela.save(
                update_fields=["status", "data_pagamento", "updated_at"]
            )
            propagate_competencia_status(parcela)
        return [correction]

    def _ensure_manual_paid(
        self,
        *,
        row: MaristelaRow,
        associado: Associado,
        referencia: date,
        target: CellTarget,
        parcela: Parcela,
        baixa: BaixaManual | None,
        valor: Decimal,
    ) -> list[dict[str, Any]]:
        corrections: list[dict[str, Any]] = []
        note = self._build_parcela_note(raw_value=target.descricao, referencia=referencia)

        if baixa is not None:
            corrections.append(
                self._build_correction(
                    row=row,
                    associado=associado,
                    contrato=parcela.ciclo.contrato,
                    referencia=referencia,
                    entity="baixa_manual",
                    entity_id=baixa.id,
                    action="soft_delete",
                    reason="manual_target_counts_for_cycle",
                    before=self._serialize_baixa(baixa),
                    after={"deleted_at": self.generated_at.isoformat()},
                )
            )
            if self.execute:
                baixa.soft_delete()

        if parcela.status == Parcela.Status.LIQUIDADA:
            return corrections

        desired_date = (
            parcela.data_pagamento
            or (baixa.data_baixa if baixa is not None else None)
            or _default_paid_date(referencia)
        )
        desired_observacao = _append_note(parcela.observacao, note)
        if (
            parcela.status == Parcela.Status.DESCONTADO
            and parcela.data_pagamento == desired_date
            and parcela.observacao == desired_observacao
        ):
            return corrections

        corrections.append(
            self._build_correction(
                row=row,
                associado=associado,
                contrato=parcela.ciclo.contrato,
                referencia=referencia,
                entity="parcela",
                entity_id=parcela.id,
                action="update_status",
                reason="sheet_manual_paid_in_cycle",
                before=self._serialize_parcela(parcela),
                after={
                    **self._serialize_parcela(parcela),
                    "status": Parcela.Status.DESCONTADO,
                    "data_pagamento": _reference_label(desired_date),
                    "observacao": desired_observacao,
                },
            )
        )
        if self.execute:
            parcela.status = Parcela.Status.DESCONTADO
            parcela.data_pagamento = desired_date
            parcela.observacao = desired_observacao
            parcela.save(
                update_fields=["status", "data_pagamento", "observacao", "updated_at"]
            )
            propagate_competencia_status(parcela)
        return corrections

    def _ensure_unpaid(
        self,
        *,
        row: MaristelaRow,
        associado: Associado,
        referencia: date,
        target: CellTarget,
        parcela: Parcela,
    ) -> list[dict[str, Any]]:
        if (
            parcela.status == Parcela.Status.NAO_DESCONTADO
            and parcela.data_pagamento is None
        ):
            return []

        correction = self._build_correction(
            row=row,
            associado=associado,
            contrato=parcela.ciclo.contrato,
            referencia=referencia,
            entity="parcela",
            entity_id=parcela.id,
            action="update_status",
            reason="sheet_unpaid",
            before=self._serialize_parcela(parcela),
            after={
                **self._serialize_parcela(parcela),
                "status": Parcela.Status.NAO_DESCONTADO,
                "data_pagamento": None,
                "observacao": parcela.observacao,
            },
        )
        if self.execute:
            parcela.status = Parcela.Status.NAO_DESCONTADO
            parcela.data_pagamento = None
            parcela.save(
                update_fields=["status", "data_pagamento", "updated_at"]
            )
            propagate_competencia_status(parcela)
        return [correction]

    def _sync_pagamentos(
        self,
        *,
        row: MaristelaRow,
        associado: Associado,
        referencia: date,
        target: CellTarget,
        pagamentos: list[PagamentoMensalidade],
        valor: Decimal,
    ) -> list[dict[str, Any]]:
        corrections: list[dict[str, Any]] = []

        if not pagamentos and target.kind != "manual":
            return corrections

        if not pagamentos and target.kind == "manual":
            correction = self._build_correction(
                row=row,
                associado=associado,
                contrato=None,
                referencia=referencia,
                entity="pagamento_mensalidade",
                entity_id=None,
                action="create",
                reason="missing_manual_payment_record",
                before={},
                after={
                    "referencia_month": _reference_label(referencia),
                    "status_code": "M",
                    "manual_status": PagamentoMensalidade.ManualStatus.PAGO,
                    "valor": str(valor),
                },
            )
            corrections.append(correction)
            if self.execute:
                PagamentoMensalidade.objects.create(
                    created_by=self.actor,
                    import_uuid=f"maristela-{referencia.strftime('%Y%m')}-{associado.id}",
                    referencia_month=referencia,
                    status_code="M",
                    matricula=associado.matricula_orgao or associado.matricula or row.matricula_planilha,
                    orgao_pagto="",
                    nome_relatorio=associado.nome_completo,
                    cpf_cnpj=associado.cpf_cnpj,
                    associado=associado,
                    valor=valor,
                    recebido_manual=valor,
                    manual_status=PagamentoMensalidade.ManualStatus.PAGO,
                    manual_paid_at=_default_manual_paid_at(referencia),
                    manual_forma_pagamento=MANUAL_PAYMENT_KIND_MARISTELA_IN_CYCLE,
                    manual_by=self.actor,
                    source_file_path=SHEET_SOURCE_PATH,
                )
            return corrections

        for pagamento in pagamentos:
            before = self._serialize_pagamento(pagamento)
            changed_fields: dict[str, Any] = {}
            if pagamento.associado_id != associado.id:
                changed_fields["associado_id"] = associado.id
            if pagamento.cpf_cnpj != associado.cpf_cnpj:
                changed_fields["cpf_cnpj"] = associado.cpf_cnpj
            if not pagamento.matricula and (associado.matricula_orgao or associado.matricula or row.matricula_planilha):
                changed_fields["matricula"] = (
                    associado.matricula_orgao or associado.matricula or row.matricula_planilha
                )
            if not pagamento.nome_relatorio:
                changed_fields["nome_relatorio"] = associado.nome_completo
            if pagamento.valor in (None, Decimal("0")) and valor:
                changed_fields["valor"] = valor

            if target.kind == "manual":
                changed_fields.update(
                    {
                        "status_code": "M",
                        "manual_status": PagamentoMensalidade.ManualStatus.PAGO,
                        "manual_paid_at": _default_manual_paid_at(referencia),
                        "manual_forma_pagamento": MANUAL_PAYMENT_KIND_MARISTELA_IN_CYCLE,
                        "manual_by_id": self.actor.id if self.actor else None,
                        "recebido_manual": valor,
                    }
                )
                if not pagamento.source_file_path:
                    changed_fields["source_file_path"] = SHEET_SOURCE_PATH
            else:
                changed_fields["status_code"] = target.pagamento_status_code
                changed_fields["manual_status"] = None
                changed_fields["manual_paid_at"] = None
                changed_fields["manual_forma_pagamento"] = ""
                changed_fields["manual_by_id"] = None
                changed_fields["recebido_manual"] = None

            after = dict(before)
            after.update(
                {
                    "status_code": changed_fields.get("status_code", before["status_code"]),
                    "manual_status": changed_fields.get("manual_status", before["manual_status"]),
                    "manual_paid_at": _reference_label(
                        changed_fields["manual_paid_at"].date()
                    )
                    if changed_fields.get("manual_paid_at")
                    else before["manual_paid_at"],
                    "manual_forma_pagamento": changed_fields.get(
                        "manual_forma_pagamento",
                        before["manual_forma_pagamento"],
                    ),
                    "recebido_manual": str(changed_fields["recebido_manual"])
                    if changed_fields.get("recebido_manual") is not None
                    else before["recebido_manual"],
                    "associado_id": changed_fields.get("associado_id", before["associado_id"]),
                    "cpf_cnpj": changed_fields.get("cpf_cnpj", before["cpf_cnpj"]),
                    "matricula": changed_fields.get("matricula", before["matricula"]),
                    "nome_relatorio": changed_fields.get(
                        "nome_relatorio",
                        before["nome_relatorio"],
                    ),
                    "valor": str(changed_fields["valor"])
                    if changed_fields.get("valor") is not None
                    else before["valor"],
                }
            )

            if before == after:
                continue

            corrections.append(
                self._build_correction(
                    row=row,
                    associado=associado,
                    contrato=None,
                    referencia=referencia,
                    entity="pagamento_mensalidade",
                    entity_id=pagamento.id,
                    action="update",
                    reason=f"sheet_{target.kind}",
                    before=before,
                    after=after,
                )
            )
            if self.execute:
                update_fields: list[str] = []
                for field_name, value in [
                    ("associado", associado if changed_fields.get("associado_id") else None),
                    ("cpf_cnpj", changed_fields.get("cpf_cnpj")),
                    ("matricula", changed_fields.get("matricula")),
                    ("nome_relatorio", changed_fields.get("nome_relatorio")),
                    ("valor", changed_fields.get("valor")),
                    ("status_code", changed_fields.get("status_code")),
                    ("manual_status", changed_fields.get("manual_status")),
                    ("manual_paid_at", changed_fields.get("manual_paid_at")),
                    ("manual_forma_pagamento", changed_fields.get("manual_forma_pagamento")),
                    ("recebido_manual", changed_fields.get("recebido_manual")),
                    ("manual_by", self.actor if "manual_by_id" in changed_fields and self.actor else None),
                    ("source_file_path", changed_fields.get("source_file_path")),
                ]:
                    if field_name == "associado":
                        if changed_fields.get("associado_id"):
                            pagamento.associado = associado
                            update_fields.append("associado")
                        continue
                    if field_name == "manual_by":
                        if "manual_by_id" in changed_fields:
                            pagamento.manual_by = self.actor if changed_fields["manual_by_id"] else None
                            update_fields.append("manual_by")
                        continue
                    if field_name in changed_fields:
                        setattr(pagamento, field_name, value)
                        update_fields.append(field_name)
                if update_fields:
                    pagamento.save(update_fields=[*sorted(set(update_fields)), "updated_at"])

        return corrections

    def _list_system_outside_sheet(
        self,
        sheet_cpfs: set[str],
        sheet_matriculas: set[str],
    ) -> list[dict[str, Any]]:
        references_by_associado = defaultdict(list)
        for associado_id, referencia in (
            Parcela.all_objects.filter(
                referencia_mes__gte=PERIOD_START,
                referencia_mes__lte=PERIOD_END,
                deleted_at__isnull=True,
            )
            .exclude(status=Parcela.Status.CANCELADO)
            .order_by("associado_id", "referencia_mes")
            .values_list("associado_id", "referencia_mes")
        ):
            references_by_associado[associado_id].append(_reference_label(referencia))

        contract_counts = defaultdict(int)
        active_contract_counts = defaultdict(int)
        for associado_id, status in Contrato.objects.values_list("associado_id", "status"):
            contract_counts[associado_id] += 1
            if status == Contrato.Status.ATIVO:
                active_contract_counts[associado_id] += 1

        rows: list[dict[str, Any]] = []
        for associado in self._all_associados:
            matricula_keys = {
                normalize_matricula(associado.matricula),
                normalize_matricula(associado.matricula_orgao),
            }
            if (
                only_digits(associado.cpf_cnpj) in sheet_cpfs
                or any(key and key in sheet_matriculas for key in matricula_keys)
            ):
                continue
            rows.append(
                {
                    "associado_id": associado.id,
                    "cpf_cnpj": associado.cpf_cnpj,
                    "nome_completo": associado.nome_completo,
                    "status": associado.status,
                    "matricula": associado.matricula,
                    "matricula_orgao": associado.matricula_orgao,
                    "total_contratos": contract_counts.get(associado.id, 0),
                    "contratos_ativos": active_contract_counts.get(associado.id, 0),
                    "referencias_no_periodo": references_by_associado.get(associado.id, []),
                }
            )
        return rows

    def _run_post_process(self) -> dict[str, Any]:
        from apps.importacao.management.commands.repair_manual_return_conflicts import (
            Command as RepairManualReturnConflictsCommand,
        )

        repair_command = RepairManualReturnConflictsCommand()
        repair_command.handle(competencia="2026-03", cpf=None, execute=True)
        return {
            "repair_manual_return_conflicts": getattr(repair_command, "summary", {}),
        }

    def _build_parcela_note(self, *, raw_value: str, referencia: date) -> str:
        return (
            f"Conciliação Maristela {self.generated_at.date().strftime('%d/%m/%Y')}: "
            f"{referencia.strftime('%m/%Y')} -> {raw_value}"
        )

    def _build_associado_note(self, status: str) -> str:
        label_map = {
            Associado.Status.ATIVO: "ativo",
            Associado.Status.INADIMPLENTE: "inadimplente",
            Associado.Status.INATIVO: "inativo",
        }
        return (
            "Status global ajustado pela conciliação Maristela "
            f"para {label_map.get(status, status)} em {self.generated_at.date().strftime('%d/%m/%Y')}."
        )

    def _build_manual_receipt_file(
        self,
        associado: Associado,
        referencia: date,
        note: str,
    ) -> ContentFile:
        content = (
            "Conciliação Maristela\n"
            f"CPF: {associado.cpf_cnpj}\n"
            f"Associado: {associado.nome_completo}\n"
            f"Competência: {referencia.strftime('%m/%Y')}\n"
            f"Observação: {note}\n"
        ).encode("utf-8")
        return ContentFile(
            content,
            name=f"conciliacao_maristela_{associado.cpf_cnpj}_{referencia.strftime('%Y%m')}.txt",
        )

    def _build_correction(
        self,
        *,
        row: MaristelaRow,
        associado: Associado,
        contrato: Contrato | None,
        referencia: date | None,
        entity: str,
        entity_id: int | None,
        action: str,
        reason: str,
        before: dict[str, Any],
        after: dict[str, Any],
    ) -> dict[str, Any]:
        return {
            "row_number": row.row_number,
            "cpf_cnpj": associado.cpf_cnpj,
            "nome_completo": associado.nome_completo,
            "associado_id": associado.id,
            "contrato_id": contrato.id if contrato is not None else None,
            "contrato_codigo": contrato.codigo if contrato is not None else "",
            "competencia": _reference_label(referencia),
            "entity": entity,
            "entity_id": entity_id,
            "action": action,
            "reason": reason,
            "before": before,
            "after": after,
        }

    def _build_exception_row(
        self,
        row: MaristelaRow,
        referencia: date,
        reason: str,
        detail: dict[str, Any],
    ) -> dict[str, Any]:
        return {
            "row_number": row.row_number,
            "cpf_cnpj": row.cpf_cnpj,
            "nome_planilha": row.nome_planilha,
            "matricula_planilha": row.matricula_planilha,
            "competencia": _reference_label(referencia),
            "reason": reason,
            "detail": detail,
        }

    @staticmethod
    def _serialize_parcela(parcela: Parcela) -> dict[str, Any]:
        return {
            "id": parcela.id,
            "contrato_id": parcela.ciclo.contrato_id,
            "contrato_codigo": getattr(parcela.ciclo.contrato, "codigo", ""),
            "status": parcela.status,
            "data_pagamento": _reference_label(parcela.data_pagamento),
            "referencia_mes": _reference_label(parcela.referencia_mes),
            "valor": str(parcela.valor),
            "observacao": parcela.observacao,
        }

    @staticmethod
    def _serialize_baixa(baixa: BaixaManual) -> dict[str, Any]:
        return {
            "id": baixa.id,
            "parcela_id": baixa.parcela_id,
            "data_baixa": _reference_label(baixa.data_baixa),
            "valor_pago": str(baixa.valor_pago),
            "observacao": baixa.observacao,
        }

    @staticmethod
    def _serialize_pagamento(pagamento: PagamentoMensalidade) -> dict[str, Any]:
        return {
            "id": pagamento.id,
            "associado_id": pagamento.associado_id,
            "cpf_cnpj": pagamento.cpf_cnpj,
            "matricula": pagamento.matricula,
            "nome_relatorio": pagamento.nome_relatorio,
            "referencia_month": _reference_label(pagamento.referencia_month),
            "status_code": pagamento.status_code,
            "valor": str(pagamento.valor) if pagamento.valor is not None else None,
            "manual_status": pagamento.manual_status,
            "manual_paid_at": (
                _reference_label(pagamento.manual_paid_at.date())
                if pagamento.manual_paid_at
                else None
            ),
            "manual_forma_pagamento": pagamento.manual_forma_pagamento,
            "recebido_manual": (
                str(pagamento.recebido_manual)
                if pagamento.recebido_manual is not None
                else None
            ),
            "source_file_path": pagamento.source_file_path,
        }


def write_maristela_reports(payload: dict[str, Any], report_dir: str | Path) -> dict[str, str]:
    target_dir = Path(report_dir).expanduser()
    target_dir.mkdir(parents=True, exist_ok=True)

    summary_name = (
        "execute_summary.json"
        if payload["summary"]["mode"] == "execute"
        else "dry_run_summary.json"
    )
    files_to_write = {
        summary_name: payload["summary"],
        "correcoes_aplicadas.json": payload["corrections"],
        "planilha_sem_match.json": payload["planilha_sem_match"],
        "excecoes_conciliacao.json": payload["excecoes_conciliacao"],
        "sistema_fora_da_planilha.json": payload["sistema_fora_da_planilha"],
        "sistema_fora_da_planilha_priorizados.json": payload[
            "sistema_fora_da_planilha_priorizados"
        ],
    }

    written_paths: dict[str, str] = {}
    for file_name, content in files_to_write.items():
        output_path = target_dir / file_name
        output_path.write_text(
            json.dumps(content, ensure_ascii=False, indent=2, default=str) + "\n",
            encoding="utf-8",
        )
        written_paths[file_name] = str(output_path)

    csv_specs = {
        "correcoes_aplicadas.csv": (payload["corrections"], CORRECTION_FIELDS),
        "planilha_sem_match.csv": (payload["planilha_sem_match"], UNMATCHED_FIELDS),
        "excecoes_conciliacao.csv": (payload["excecoes_conciliacao"], EXCEPTION_FIELDS),
        "sistema_fora_da_planilha.csv": (
            payload["sistema_fora_da_planilha"],
            SYSTEM_ONLY_FIELDS,
        ),
        "sistema_fora_da_planilha_priorizados.csv": (
            payload["sistema_fora_da_planilha_priorizados"],
            SYSTEM_ONLY_FIELDS,
        ),
    }
    for file_name, (rows, base_fields) in csv_specs.items():
        output_path = target_dir / file_name
        _write_csv(output_path, rows, base_fields)
        written_paths[file_name] = str(output_path)

    return written_paths


def _write_csv(path: Path, rows: list[dict[str, Any]], base_fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(base_fields)
    extra_fields = sorted(
        {
            key
            for row in rows
            for key in row.keys()
            if key not in fields
        }
    )
    fields.extend(extra_fields)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            serialized = {}
            for key in fields:
                value = row.get(key)
                if isinstance(value, (dict, list)):
                    serialized[key] = json.dumps(value, ensure_ascii=False, default=str)
                else:
                    serialized[key] = value
            writer.writerow(serialized)


__all__ = [
    "MaristelaReconciliationRunner",
    "PERIOD_REFERENCES",
    "write_maristela_reports",
    "_default_report_dir",
]
